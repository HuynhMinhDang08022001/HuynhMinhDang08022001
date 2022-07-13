from email import message
from turtle import textinput
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color, numbers
import os
from struct import pack
import tkinter as tk
from PIL import Image, ImageTk
from tkinter.filedialog import askopenfile
from tkinter import filedialog
from tkinter import HORIZONTAL, messagebox
from tkinter import ttk
import tkinter.ttk
import time
import string
from tkinter.filedialog import askopenfile


try:
    os.system('TASKKILL /F /IM EXCEL.exe')
    os.remove("TC_RF.xlsx")
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "TC_RF"
    wb2.save("TC_RF.xlsx")
    os.close("TC_RF.xlsx")
except:
    try:
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "TC_RF"
        wb2.save("TC_RF.xlsx")
        os.system('TASKKILL /F /IM EXCEL.exe')
    except OSError:
        print('Failed creating the file')
    else:
        print('File created')

def create_value_file():
    try:
        fd = os.open("RFvalue.xlsx", os.O_RDWR)
        os.close(fd)
    except OSError:
        z = 0
        x = 0
        wb5 = Workbook()
        ws5 = wb5.active
        sheet = wb5.worksheets[0]
        noneFill = PatternFill(start_color='00FFFFFF',
                                end_color='00FFFFFF',
                                fill_type='solid'
                                )
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))
        font_text = Font(name="Calibri", size=14, color='00FFFFFF', bold=True)
        alignment = Alignment(horizontal='center', vertical='center')
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
        sheet.column_dimensions['A'].alignment = alignment
        sheet.column_dimensions['B'].alignment = alignment
        sheet.column_dimensions['C'].alignment = alignment
        sheet.column_dimensions['D'].alignment = alignment
        sheet.column_dimensions['F'].alignment = alignment
        sheet.column_dimensions['G'].alignment = alignment

        ws5.title = "RFvalue_baseSW"
        ws5.append(['DID', 'Description', 'Length (Byte)', 'Value'])

        ws5['F1'] = 'BaseSW Name'
        cell_header = ws5.cell(1, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Ticket BaseSW'])
        ws5['F2'] = 'Ticket BaseSW'
        cell_header = ws5.cell(2, 6)
        cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Variant BaseSW'])
        ws5['F3'] = 'Variant BaseSW'
        cell_header = ws5.cell(3, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'DID check variant BaseSW'])
        ws5['F4'] = 'DID check variant BaseSW'
        cell_header = ws5.cell(4, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        for col in range(1, 7):
            cell_header = ws5.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment

        cell_header = ws5.cell(1, 5)
        cell_header.fill = noneFill
        cell_header.border = border

        wb5.create_sheet("RFvalue_latestSW")
        ws5 = wb5['RFvalue_latestSW']
        sheet2 = wb5.worksheets[1]
        sheet2.column_dimensions['B'].width = 50
        sheet2.column_dimensions['C'].width = 20
        sheet2.column_dimensions['D'].width = 30
        sheet2.column_dimensions['F'].width = 30
        sheet2.column_dimensions['A'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['B'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['C'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['D'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['F'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['G'].number_format = numbers.FORMAT_TEXT
        sheet2.column_dimensions['A'].alignment = alignment
        sheet2.column_dimensions['B'].alignment = alignment
        sheet2.column_dimensions['C'].alignment = alignment
        sheet2.column_dimensions['D'].alignment = alignment
        sheet2.column_dimensions['F'].alignment = alignment
        sheet2.column_dimensions['G'].alignment = alignment
        # column_count2 = sheet2.max_column
        ws5.append(['DID', 'Description', 'Length (Byte)', 'Value'])

        ws5['F1'] = 'LatestSW Name'
        cell_header = ws5.cell(1, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Ticket LatestSW'])
        ws5['F2'] = 'Ticket LatestSW'
        cell_header = ws5.cell(2, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'Variant LatestSW'])
        ws5['F3'] = 'Variant LatestSW'
        cell_header = ws5.cell(3, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        # ws5.append(['', '', '', '', '', 'DID check variant LatestSW'])
        ws5['F4'] = 'DID check variant LatestSW'
        cell_header = ws5.cell(4, 6)
        cell_header.fill = PatternFill(
            start_color='000066CC', end_color='000066CC', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
        cell_header.alignment = alignment
        for col in range(1, 7):
            cell_header = ws5.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            cell_header.alignment = alignment


        cell_header = ws5.cell(1, 5)
        cell_header.fill = noneFill
        cell_header.border = border

        wb5.save("RFvalue.xlsx")
        os.system('TASKKILL /F /IM EXCEL.exe')
        messagebox.showinfo("COMPLETE", "File RFvalue.xlsx has been created in the same folder tool successfully, Please fill all value")
        print('tao thanh cong')
        # Mo mot file
        # path = "/py/"
        # dirs = os.listdir(path)
        
        # # Lenh de in tat ca file va thu muc
        # for file in dirs:
        #     print(file)




check_DID = 0

def DID_baseSW(ws2, wb, id, number1, number2, number3, number4, direct):
    if direct == '':
        wb = load_workbook('RFvalue.xlsx')
    else:
        wb = load_workbook(str(direct))
    ws = wb.active
    ws = wb['RFvalue_baseSW']
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            baseSW = ws[char + str(row)].value
    
    i = 0
    o = 2
    j = 3

    k = 1
    number4 += 1

    i = 0
    count_string_number = 0
    hexvalue_baseSW = ""
    while k < row_count:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID_baseSW = ws[char + str(row)].value
                row_list_DID_baseSW_lowercase = str(
                    row_list_DID_baseSW).lower()

        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name_baseSW = ws[char + str(row)].value
                if str(row_list_DID_baseSW) == "None":
                    messagebox.showerror("ERROR", "No DID, Please add DID")
                    break
                # print(ws[char + str(row)].value)
        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_length_byte_baseSW = ws[char + str(row)].value

        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_values_baseSW = ws[char + str(row)].value
                
        id += 1
        # check lenghth byte
        count_hexvalue_baseSW = 0
        hexvalue_baseSW = ""
        length_byte = 0
        if str(row_list_values_baseSW) == 'None':
            hexvalue_baseSW = str(
                ".{" + str(row_list_length_byte_baseSW) + "}")
            # print('dung')
        else:
            # change ascii sang hex value
            for i in str(row_list_values_baseSW):
                hexvalue_baseSW += hex(ord(i))[2:]
            # print(hexvalue_baseSW)
            count_hexvalue_baseSW = len(hexvalue_baseSW)
            count_hexvalue_baseSW = int(count_hexvalue_baseSW) // 2
            # print(count_hexvalue_baseSW)
            # print(type(row_list_length_byte_baseSW))
            if str(count_hexvalue_baseSW) < row_list_length_byte_baseSW:
                length_byte = (int(row_list_length_byte_baseSW) -
                               count_hexvalue_baseSW) ** 2
                hexvalue_baseSW = hexvalue_baseSW.lower()
                hexvalue_baseSW = str(
                    hexvalue_baseSW + ".{" + str(length_byte) + "}")
            print(hexvalue_baseSW)
        if str(row_list_name_baseSW) == "None":
            row_list_name_baseSW = ""
        ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_baseSW) + ' ' + str(row_list_name_baseSW), 'To check value of the DID ' + str(row_list_DID_baseSW), '1) Send service 0x22 to the camera for the DID ' +
                    str(row_list_DID_baseSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_baseSW_lowercase) + ','+'62'+str(row_list_DID_baseSW_lowercase) + '.*' + str(hexvalue_baseSW) + '.*' + ', Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

        number4 += 1
        o += 1
        j += 1
        k += 1

    return id


def DID_latestSW(ws3, wb3, id, number1, number2, number3, number4, direct):
    if direct == '':
        wb3 = load_workbook('RFvalue.xlsx')
    else:
        wb3 = load_workbook(str(direct))
        
    ws3 = wb3 .active
    ws3 = wb3['RFvalue_latestSW']
    sheet2 = wb3.worksheets[1]
    row_count2 = sheet2.max_row
    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            latestSW = ws3[char + str(row)].value

    i = 0
    o = 2
    j = 3
    k = 1
    number4 += 1

    i = 0

    while k < row_count2:
        for row in range(o, j):
            for col in range(1, 2):
                char = get_column_letter(col)
                row_list_DID_latestSW = ws3[char + str(row)].value
                row_list_DID_latestSW_lowercase = str(
                    row_list_DID_latestSW).lower()
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws3[char + str(row)].value)
        for row in range(o, j):
            for col in range(2, 3):
                char = get_column_letter(col)
                row_list_name_latestSW = ws3[char + str(row)].value
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws3[char + str(row)].value)
        for row in range(o, j):
            for col in range(3, 4):
                char = get_column_letter(col)
                row_list_length_byte_latestSW = ws3[char + str(row)].value
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws[char + str(row)].value)
                # print(row_list_length_byte_latestSW)
        for row in range(o, j):
            for col in range(4, 5):
                char = get_column_letter(col)
                row_list_values_latestSW = ws3[char + str(row)].value
                if str(row_list_DID_latestSW) == "None":
                    return id
                # print(ws3[char + str(row)].value)
        if str(row_list_DID_latestSW) != "None":
            id += 1
            # check lenghth byte
            hexvalue_latestSW = ""
            # length_byte = ""
            if str(row_list_values_latestSW) == 'None':
                hexvalue_latestSW = str(
                    ".{" + str(row_list_length_byte_latestSW) + "}")
                # print('dung')
            else:
                # change ascii sang hex value
                for i in str(row_list_values_latestSW):
                    hexvalue_latestSW += hex(ord(i))[2:]
                count_hexvalue_latestSW = len(hexvalue_latestSW)
                count_hexvalue_latestSW = int(count_hexvalue_latestSW) // 2
                # print(count_hexvalue_latestSW)
                # print(type(row_list_length_byte_latestSW))
                if str(count_hexvalue_latestSW) < row_list_length_byte_latestSW:
                    length_byte = (int(row_list_length_byte_latestSW) -
                                   count_hexvalue_latestSW) ** 2
                    hexvalue_latestSW = hexvalue_latestSW.lower()
                    hexvalue_latestSW = str(
                        hexvalue_latestSW + ".{" + str(length_byte) + "}")
                # print(hexvalue_latestSW)
            if str(row_list_name_latestSW) == "None":
                row_list_name_latestSW = ""
            ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' ' + str(row_list_DID_latestSW) + ' ' + str(row_list_name_latestSW), 'To check value of the DID ' + str(row_list_DID_latestSW), '1) Send service 0x22 to the camera for the DID ' +
                        str(row_list_DID_latestSW) + ' using physical addressing', '1) -', '1) RequestResponse(' + '22' + str(row_list_DID_latestSW_lowercase) + ','+'62'+str(row_list_DID_latestSW_lowercase) + '.*' + str(hexvalue_latestSW) + '.*' + ', Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
            
            
            
            number4 += 1
            o += 1
            j += 1
            k += 1
        else:
            return id
    return id


def variant_base_sw(id, number1, number2, number3, number4, tasks):
    direct = Input_path_text.get()
    if direct == '':
        wb = load_workbook('RFvalue.xlsx')
    else:
        wb = load_workbook(direct)
    ws = wb.active
    ws = wb['RFvalue_baseSW']

    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            baseSW = ws[char + str(row)].value
            # print(baseSW)

    for row in range(3, 4):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_Variant_BaseSW = ws[char + str(row)].value
            row_Variant_BaseSW_lowercase = str(row_Variant_BaseSW).lower()
            # print(row_Variant_BaseSW)
    for row in range(4, 5):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_DID_check_variant_BaseSW = ws[char + str(row)].value
            row_DID_check_variant_BaseSW_lowercase = str(
                row_DID_check_variant_BaseSW).lower()
            # print(row_DID_check_variant_BaseSW)

    if str(row_Variant_BaseSW) != "None" and str(row_DID_check_variant_BaseSW) != "None":
        ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                    '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e' + str(row_Variant_BaseSW_lowercase) + ', 6e' + str(row_Variant_BaseSW_lowercase) + ', Equal)\n7) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62' + str(row_Variant_BaseSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', baseSW, ''])
    else:
        ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)' , 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -',
                    '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(22' + str(row_DID_check_variant_BaseSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    
    return id


def variant_latest_sw(id, number1, number2, number3, number4, tasks):
    direct = Input_path_text.get()
    if direct == '':
        wb3 = load_workbook('RFvalue.xlsx')
    else:
        wb3 = load_workbook(str(direct))
    ws3 = wb3.active
    ws3 = wb3['RFvalue_latestSW']
    sheet2 = wb3.worksheets[1]
    row_count = sheet2.max_row
    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            latestSW = ws3[char + str(row)].value
            # print(latestSW)

    for row in range(3, 4):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_Variant_LatestSW = ws3[char + str(row)].value
            row_Variant_LatestSW_lowercase = str(row_Variant_LatestSW).lower()
            # print(row_Variant_LatestSW)
    for row in range(4, 5):
        for col in range(7, 8):
            char = get_column_letter(col)
            row_DID_check_variant_LatestSW = ws3[char + str(row)].value
            row_DID_check_variant_LatestSW_lowercase = str(row_DID_check_variant_LatestSW).lower()
            # print(row_DID_check_variant_LatestSW)

    if str(row_Variant_LatestSW) != "None" and str(row_DID_check_variant_LatestSW) != "None":
        ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Select_and_check_variant', 'To Select and check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Select variant\n7) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -\n7) -',
                '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(2e'+str(row_Variant_LatestSW_lowercase) + ', 6e'+str(row_Variant_LatestSW_lowercase) + ', Equal)\n7) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62' + str(row_Variant_LatestSW_lowercase) + ', Equal)', 'Automated Testcase', 'implemented', latestSW, ''])
    else:
        ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Check_variant (Variant is default)', 'To check variant', '1) Tester Present is ON\n2) Change to Extended session with Service 0x10 03\n3) Security unlock ON\n4) wait\n5) Security unlock OFF\n6) Check variant', '1) -\n2) -\n3) -\n4) -\n5) -\n6) -',
                    '1) envvar(EnvTesterPresentOnOff(1;0))\n2) RequestResponse(1003, 5003.*, Regexp)\n3) envvar(EnvLogInLevel1_1(0;0))\n4) wait(1000)\n5) envvar(EnvLogInLevel1_1(0;0))\n6) RequestResponse(22' + str(row_DID_check_variant_LatestSW_lowercase) + ', 62.*, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    return id
# -------------------------------------------------------------------------------------------------------------

# SOURCE


def run_program():
    run_btn_text.set("Loading...")

    direct = Input_path_text.get()
    if direct == '':
        wb = load_workbook('RFvalue.xlsx')
        wb3 = load_workbook('RFvalue.xlsx')
    else:
        wb = load_workbook(direct)
        wb3 = load_workbook(direct)
    wb3 = load_workbook(direct)
    ws = wb .active
    ws = wb['RFvalue_baseSW']
    ws3 = wb3['RFvalue_latestSW']
    sheet2 = wb.worksheets[1]
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    row_count2 = sheet2.max_row
    tasks = row_count2 + row_count + 192
    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            baseSW = ws[char + str(row)].value
    for row in range(2, 3):
        for col in range(7, 8):
            char = get_column_letter(col)
            ticket_baseSW = ws[char + str(row)].value
            # print(ticket_baseSW)
    for row in range(1, 2):
        for col in range(7, 8):
            char = get_column_letter(col)
            latestSW = ws3[char + str(row)].value
    for row in range(2, 3):
        for col in range(7, 8):
            char = get_column_letter(col)
            ticket_latestSW = ws3[char + str(row)].value
    
    # script begin
    id = 2
    number1 = 1
    number2 = 1
    number3 = 1
    number4 = 1
    border = Border(left=Side(border_style='thin', color='000000'),right = Side(border_style='thin', color='000000'),top = Side(border_style='thin', color='000000'),bottom = Side(border_style='thin', color='000000'))
    font_text_header = Font(name="Calibri", size=13, color='00FFFFFF', bold=True)
    font_text = Font(name="Calibri", size=11, color='00000000', bold=False)
    alignment = Alignment(horizontal='center', vertical='center')
    ws2.append(['ID', 'XXX Component',  'Test Description', 'Test Steps',  'Test Response','Teststep keywords', 'ObjectType', 'TestStatus', 'Project', 'TestResult'])
    for col in range(1, 11):
            cell_header = ws2.cell(1, col)
            # used hex code for red color
            cell_header.fill = PatternFill(start_color='000066CC', end_color='000066CC', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text_header
            # cell_header.alignment = alignment
    ws2.append(['ID_'+str(id),  '1 REFFLASH', '', '', '', '', 'Test group', '', '', ''])
    for col in range(1, 11):
            cell_header = ws2.cell(2, col)
            # used hex code for red color
            cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
            # cell_header.alignment = alignment
    # ------------------------------------------------------------------------------------------------------
    # BEGIN TEST CASE 1
    # TEST CASE 1 base SW to latestSW M3
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) +
                ' Base SW to Latest SW M3', '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) +
                ' Flash Base SW via UART', '', '', '', '', 'Test group', '', '', ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' +
                str(number3) + ' Flash SW', '', '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash' + str(baseSW), 'Detail information is mentioned in the ticket: ' +
                str(ticket_baseSW), 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', str(baseSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_base_sw(id, number1, number2, number3, number4, tasks)

    # id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    direct = Input_path_text.get()
    id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    # print(id)
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # Step2 FLASH LATEST_SW M3 VIA Xflash TOOLS

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    sheet = wb.worksheets[1]
    row_count = sheet.max_row
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+ticket_latestSW,
                'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', latestSW, ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*1, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*1, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')

    # END TEST CASE 1

    # ------------------------------------------------------------------------------------------------------
    # # BEGIN TEST CASE 2
    # # TEST CASE 2 base SW to latestSW M5
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' Base SW to Latest SW M3',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH BASE_SW VIA UART script
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash Base SW via UART',
                '', '', '', '', 'Test group', '', '', ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])

    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash' + str(baseSW), 'Detail information is mentioned in the ticket: ' + str(ticket_baseSW),
                'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', str(baseSW), ''])

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(
            start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_base_sw(id, number1, number2, number3, number4, tasks)

    # id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    # number4 += 1
    direct = Input_path_text.get()
    id = DID_baseSW(ws2, wb, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*0, Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', baseSW, ''])

    # # ws2.title = "TC_RF"
    # # wb2.save('TC_RF.xlsx')

    # # Step2 FLASH LATEST_SW M5 1st VIA Xflash TOOLS

    # # # Reflash latest SW M5 via xflash tool

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M5 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+ticket_latestSW,
                'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*1, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*1, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(row, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*5,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # # Step2 FLASH LATEST_SW M3 2nd VIA Xflash TOOLS

    # # # Reflash latest SW M3 via xflash tool

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Latest SW M3 via X-Flash 2nd',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+ticket_latestSW,
                'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*2, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*2, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # # ws2.title = "TC_RF"
    # # wb2.save('TC_RF.xlsx')

    # # END TEST CASE 2

    # # ------------------------------------------------------------------------------------------------------

    # # BEGIN TEST CASE 3
    # # TEST CASE 3 latest SW to DummySW M3
    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' latest SW to Dummy SW M3',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH latest_SW VIA UART script

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash latest SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])
    
    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash'+latestSW, 'Detail information is mentioned in the ticket: '+ticket_latestSW,
                'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    id = DID_latestSW(ws2, wb, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # # ws2.title = "TC_RF"
    # # wb2.save('TC_RF.xlsx')

    # # Step2 FLASH DUMMY_SW M3 VIA Xflash TOOLS

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M3 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+ticket_latestSW,
                'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])
    
    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct)
    # print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*1, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*1, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # # ws2.title = "TC_RF"
    # # wb2.save('TC_RF.xlsx')

    # # END TEST CASE 3

    # # ------------------------------------------------------------------------------------------------------
    # # BEGIN TEST CASE 4
    # # TEST CASE 4 latest SW to DummySW M5
    number1 += 1
    number2 = 1
    number3 = 1
    number4 = 1
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + ' latest SW to Dummy SW M5',
                '', '', '', '', 'Test group', '', '', ''])

    # # step 1 FLASH latest_SW VIA UART script

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Flash latest SW via UART',
                '', '', '', '', 'Test group', '', '', ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Flash SW', '',
                '', '', '', 'Test group', '', '', ''])
    
    for row in range(id - 2, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.fill = PatternFill(
                start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
            cell_header.border = border
            cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' UART flash'+latestSW, 'Detail information is mentioned in the ticket: '+ticket_latestSW,
                'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()

    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    # number4 += 1
    direct = Input_path_text.get()
    id = DID_latestSW(ws2, wb, id, number1, number2, number3, number4, direct)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*0, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # ws2.title = "TC_RF"
    # wb2.save('TC_RF.xlsx')
    
# --------------------------------------------------------------------------------------------------------
    # Step2 FLASH DUMMY_SW M5 1st VIA Xflash TOOLS
    # # Reflash Dummy SW M5 via xflash tool

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M5 via X-Flash 1st',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+ticket_latestSW,
                'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct)
    # print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*1, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*1, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*5,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # Step2 FLASH DUMMY_SW M3 2nd VIA Xflash TOOLS

    # # Reflash Dummy SW M3 via xflash tool

    ws3 = wb .active
    ws3 = wb3['RFvalue_latestSW']

    # number = df.shape[0]
    # print(number)
    sheet = wb.worksheets[1]
    row_count = sheet.max_row

    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number2 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + ' Re-Flash Dummy SW M3 via X-Flash 2nd',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 = 1
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' Xflash '+latestSW, 'Detail information is mentioned in the ticket: '+ticket_latestSW,
                'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Screen shot the successful flash proccess', 'Manual Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Variant and Software  Identification',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    id = variant_latest_sw(id, number1, number2, number3, number4, tasks)

    direct = Input_path_text.get()
    # DID_latestSW(ws3, wb, id, number1, number2, number3, number4)
    id = DID_latestSW(ws3, wb, id, number1, number2, number3, number4, direct)
    # print(id)

    # programing couter
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' Programming Counter and Programming Attempt Counter',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0200_ProgrammingCounter', 'To check value of the DID 0200', '1) Send service 0x22 to the camera for the DID 0200 using physical addressing',
                '1) -', '1) RequestResponse(220200, 620200.*2, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 0201_ProgrammingAttemptCounter', 'To check value of the DID 0201', '1) Send service 0x22 to the camera for the DID 0201 using physical addressing',
                '1) -', '1) RequestResponse(220201, 620201.*2, Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])

    # RBEOL read
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number3 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + ' DID in RBEOL',
                '', '', '', '', 'Test group', '', '', ''])

    for col in range(1, 11):
        cell_header = ws2.cell(id, col)
        # used hex code for red color
        cell_header.fill = PatternFill(start_color='0000CCFF', end_color='0000CCFF', fill_type="solid")
        cell_header.border = border
        cell_header.font = font_text
    
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 = 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1E0', 'To check value of the DID F1E0', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1E0 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1e0,62f1e0.*3,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' F1DD', 'To check value of the DID F1DD', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID F1DD using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(22f1dd,62f1dd.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4255_RBEOL_DID', 'To check value of the DID 4255', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224255,624255.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    id += 1
    bar['value'] += 1
    percent.set(str((id//tasks)*100)+"%")
    app.update_idletasks()
    number4 += 1
    ws2.append(['ID_'+str(id),  '1.' + str(number1) + '.' + str(number2) + '.' + str(number3) + '.' + str(number4) + ' 4259_RBEOLDID', 'To check value of the DID 4259', '1) Access to RBEOL\n2) Unlock RBEOL\n3) Wait 5s\n4) Send service 0x22 to the camera for the DID 4255 using physical addressing', '1) -\n2) -\n3) -\n4) -',
                '1) envvar(EnvRBEOL(1;1000), EnvRBEOL(0;1000))\n2) envvar(Env_MPC3_EOL_unlock(1;1000), Env_MPC3_EOL_unlock(0;1000))\n3) wait(5000)\n4) RequestResponse(224259,624259.*,Regexp)', 'Automated Testcase', 'implemented', latestSW, ''])
    for row in range(id - id + 1, id + 1):
        for col in range(1, 11):
            cell_header = ws2.cell(row, col)
            # used hex code for red color
            cell_header.border = border
    
    locate_save = Output_path_text.get()
    ws2.title = "TC_RF"
    print(locate_save)
    if locate_save == '':
        # print('Ok') 
        wb2.save('TC_RF.xlsx')
    else:
        wb2.save(locate_save +'/'+'TC_RF.xlsx')
        
    if locate_save == '':
        wb6 = load_workbook(direct)
    else:
        wb6 = load_workbook(locate_save +'/'+'TC_RF.xlsx')
        
    ws6 = wb6['TC_RF']
    sheet3 = wb6.worksheets[0]
    sheet3.column_dimensions['A'].width = 30
    sheet3.column_dimensions['B'].width = 30
    sheet3.column_dimensions['C'].width = 30
    sheet3.column_dimensions['D'].width = 30
    sheet3.column_dimensions['E'].width = 30
    sheet3.column_dimensions['F'].width = 30
    sheet3.column_dimensions['G'].width = 30
    sheet3.column_dimensions['H'].width = 30
    sheet3.column_dimensions['I'].width = 30
    sheet3.column_dimensions['J'].width = 30
    
    if locate_save == '': 
        wb6.save('TC_RF.xlsx')
    else:
        wb6.save(locate_save +'/'+'TC_RF.xlsx')
    # END TEST CASE 4
    
    run_btn_text.set("DONE")
    tkinter.messagebox.showinfo("GREAT!", "Test case RFlash tool created successfully")


def start_program():
    direct = Input_path_text.get()
    if direct == '':
        wb = load_workbook('RFvalue.xlsx')
        wb3 = load_workbook('RFvalue.xlsx')
    else:
        wb = load_workbook(direct)
        wb3 = load_workbook(direct)
    ws = wb.active
    ws = wb['RFvalue_baseSW']
    ws3 = wb3.active
    ws3 = wb3['RFvalue_latestSW']
    sheet2 = wb.worksheets[1]
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    for row in range(13, 14):
        for col in range(2, 3):
            char = get_column_letter(col)
            baseSW = ws[char + str(row)].value
            # print(baseSW)
    for row in range(14, 15):
        for col in range(2, 3):
            char = get_column_letter(col)
            ticket_baseSW = ws[char + str(row)].value
            # print(ticket_baseSW)
    for row in range(15, 16):
        for col in range(2, 3):
            char = get_column_letter(col)
            row_Variant_BaseSW = ws[char + str(row)].value
            row_Variant_BaseSW_lowercase = str(row_Variant_BaseSW).lower()
            # print(row_Variant_BaseSW)
    for row in range(17, 18):
        for col in range(2, 3):
            char = get_column_letter(col)
            row_DID_check_variant_BaseSW = ws[char + str(row)].value
            row_DID_check_variant_BaseSW_lowercase = str(
                row_DID_check_variant_BaseSW).lower()
            # print(row_DID_check_variant_BaseSW)
    for row in range(13, 14):
        for col in range(2, 3):
            char = get_column_letter(col)
            latestSW = ws3[char + str(row)].value
            # print(latestSW)
    for row in range(14, 15):
        for col in range(2, 3):
            char = get_column_letter(col)
            ticket_latestSW = ws3[char + str(row)].value
            # print(ticket_latestSW)
    for row in range(15, 16):
        for col in range(2, 3):
            char = get_column_letter(col)
            row_Variant_LatestSW = ws3[char + str(row)].value
            row_Variant_LatestSW_lowercase = str(row_Variant_LatestSW).lower()
            # print(row_Variant_LatestSW)
    for row in range(17, 18):
        for col in range(2, 3):
            char = get_column_letter(col)
            row_DID_check_variant_LatestSW = ws[char + str(row)].value

    for row in range(2, 3):
        for col in range(1, 2):
            char = get_column_letter(col)
            row_list_DID_baseSW = ws[char + str(row)].value
            row_list_DID_latestSW = ws3[char + str(row)].value
                
    if str(row_list_DID_baseSW) == "None":
        messagebox.showerror("ERROR", "No DID in row 1 at sheet RFvalue_baseSW, Please add DID")
        
    if str(row_list_DID_latestSW) == "None":
        messagebox.showerror("ERROR", "No DID in row 1 at sheet RFvalue_latestSW, Please add DID")
        
    if baseSW == "":
        messagebox.showerror("ERROR", "BaseSW Name is invalid")
        
    if ticket_baseSW == "":
        messagebox.showerror("ERROR", "ticket_baseSW Name is invalid")
        
    if latestSW == "":
        messagebox.showerror("ERROR", "latestSW Name is invalid")
        
    if ticket_latestSW == "":
        messagebox.showerror("ERROR", "ticket_latestSW Name is invalid")
        
    if baseSW != "" and ticket_baseSW != "" and latestSW != "" and ticket_latestSW != "" and str(row_list_DID_baseSW) != "None" and str(row_list_DID_latestSW) != "None":
        print('run_program')
        run_program()


# app front end

app = tk.Tk()

app.title('Reflash Tool')
app.geometry('700x420')


# logo
logo = Image.open('background_img.jpg')
logo = ImageTk.PhotoImage(logo)


# Creat canvas

canvas = tk.Canvas(app, width=700, height=420)

canvas.pack(fill='both', expand=True)
canvas.create_image(0, 0, image=logo, anchor='nw')

create_value_file()

def open_file():
    browse_input_path_text.set("loading...")
    file_path = askopenfile(parent=app, mode='rb', title="Choose location take file", filetype=[
                            ("excel file", ".xlsx")])
    print("Original string: " + str(file_path))

    result_str = ""
    final_str = ""
    for i in range(0, len(str(file_path))):
        if i >= 26:
            result_str = result_str + str(file_path)[i]
    reverse_str = result_str[::-1]
    for i in range(0, len(reverse_str)):
        if i >= 2:
            final_str = final_str + reverse_str[i]
    complete_str = final_str[::-1]
    print(type(complete_str))
    print(complete_str)
    if file_path:
        Input_path_text.set(str(complete_str))
        browse_input_path_text.set("Browse")
    return complete_str


def save_file():
    # print("is this working??")
    browse_output_path_text.set("loading...")
    file_path2 = filedialog.askdirectory()
    print(file_path2)
    if file_path2:
        Output_path_text.set(str(file_path2))
        browse_output_path_text.set("Browse")
    return str(file_path2)

# WELCOME
# Welcome = tk.StringVar()
# welcome_window = canvas.create_text(
#     360, 100, text='WELCOME!', font=('helvetica', 50), fill="Black")


file_path = ""

# instruction
instruction = tk.Label(app, text="Welcome to ReFlash tool create by dev Huynh Minh Dang", font=("helvetica", 14))
instruction_window = canvas.create_window(100, 387, anchor="nw", window=instruction)

version_app = canvas.create_text(660, 403, text="R1.1.1", font=("helvetica", 12),fill="Black")
# version_app_window = canvas.create_window(200, 387, anchor="nw", window=version_app)

# Part Base SW
Input_path_text = tk.StringVar()
Input_path_label = canvas.create_text(
    80, 50, text='Input path', font=('bold', 14), fill="Black")
Input_path_entry = tk.Entry(
    app, textvariable=Input_path_text, font='large_font', width=50)
Input_path_entry_window = canvas.create_window(
    10, 70, anchor="nw", window=Input_path_entry)


Output_path_text = tk.StringVar()
Output_path_label = canvas.create_text(
    80, 130, text='Output path', font=('bold', 14), fill="Black")
Output_path_entry = tk.Entry(
    app, textvariable=Output_path_text, font='large_font', width=50)
Output_path_entry_window = canvas.create_window(
    10, 150, anchor="nw", window=Output_path_entry)


# Buttons
run_btn_text = tk.StringVar()
run_btn = tk.Button(app, textvariable=run_btn_text,
                    command=start_program, font="Raleway", width=15)
run_btn_text.set("RUN")

run_btn_window = canvas.create_window(270, 190, anchor="nw", window=run_btn)

# browse button
browse_input_path_text = tk.StringVar()
browse_btn_input_path = tk.Button(app, textvariable=browse_input_path_text,
                                    command=lambda: open_file(), font="Raleway", width=7, height=1)
browse_input_path_text.set("Browse")
browse_btn_input_path = canvas.create_window(
    475, 65, anchor="nw", window=browse_btn_input_path)

# browse button save file
browse_output_path_text = tk.StringVar()
browse_btn_output_path = tk.Button(app, textvariable=browse_output_path_text,
                                    command=lambda: save_file(), font="Raleway", width=7, height=1)
browse_output_path_text.set("Browse")
browse_output_path_window = canvas.create_window(
    475, 145, anchor="nw", window=browse_btn_output_path)

# progress bar
bar = ttk.Progressbar(app, orient=HORIZONTAL, length=600, mode='determinate')
bar_window = canvas.create_window(40, 250, anchor="nw", window=bar)

percent = tk.StringVar()
percentLabel = tk.Label(app, textvariable=percent)
percentLabel_window = canvas.create_window(330, 300, anchor="nw", window=percentLabel)

# resize
# app.resizable(True,True)

# my_sizegrip = ttk.Sizegrip(app)


# def resizer(e):
#     global logo1, resize_logo, new_logo, instruction
#     logo1 = Image.open('background_img.jpg')
#     resize_logo = logo1.resize((e.width, e.height), Image.ANTIALIAS)
#     new_logo = ImageTk.PhotoImage(resize_logo)
#     canvas.create_image(0, 0, image=new_logo, anchor='nw')

#     print(e.width)
#     size  = e.width /10
#     instruction = tk.Label(app, text="Welcome to ReFlash tool create by dev Huynh Minh Dang", font=("helvetica", int(size)))
    # instruction_window = canvas.create_window(100, 387, anchor="nw", window=instruction)
    # run_btn_text = tk.StringVar()
    # run_btn1 = tk.Button(app, textvariable=run_btn_text,
    #                     command=start_program, font="Raleway", width=15)
    # size  = e.width /10
    # run_btn.config(font=("helvetica", int(size)))
    # run_btn_text.set("RUN")

    # run_btn_window = canvas.create_window(
    #     270, 190, anchor="nw", window=run_btn)

# app.bind('<Configure>', resizer)

# Start program
app.mainloop()
